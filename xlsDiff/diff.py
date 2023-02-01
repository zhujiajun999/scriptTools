import openpyxl
import warnings
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
import os

warnings.filterwarnings('ignore')
os.system("rm -f ~/PyPject/{title}")

# 加载Excel
data_all = openpyxl.load_workbook("emoji全库v2.xlsx")  # 加载Excel文件
data_all_sheet1 = data_all[data_all.sheetnames[0]]  # 获取第一张数据表
data_ban = openpyxl.load_workbook("Win-屏蔽Emoji.xlsx")
data_ban_sheet1 = data_ban[data_ban.sheetnames[0]]

data_all_list = []
data_ban_list = []


def Textcolor(file_name, title):
    for row in range(2, data_ban_sheet1.max_row):
        banWord = data_ban_sheet1.cell(row=row, column=1).value
        for row2 in range(2, data_all_sheet1.max_row):
            allWord = data_all_sheet1.cell(row=row2, column=2).value
            # print(banWord)
            # print(allWord)
            # print("================")
            if banWord == allWord:
                Color = ['ffc7ce', '9c0006']  # 红
                fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 红色
                font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])  # 设置字体样式
                data_all_sheet1.cell(row=row2, column=2, value="").fill = fille
                data_all_sheet1.cell(row=row2, column=2, value=banWord).font = font
                data_all_sheet1.cell(row=row2, column=3, value="").fill = fille
                data_all_sheet1.cell(row=row2, column=3, value="屏蔽").font = font

    data_all.save(file_name)  # 保存excel
    data_all.close()


file_name, title = 's.xlsx', '处理'
Textcolor(file_name, title)

# # ban emoji unicode
# for row in range(2, data_ban_sheet1.max_row):
#     data_ban_dict = {}
#     banunicode = data_ban_sheet1.cell(row=row, column=1).value
#     # data_ban_dict['屏蔽Emoji'] = banunicode
#     data_ban_list.append(banunicode)

# # all emoji unicode
# for row in range(2, data_all_sheet1.max_row):
#     data_raw_dict = {}
#     unicode = data_all_sheet1.cell(row=row, column=2).value
#     # data_raw_dict['Unicode编码'] = unicode
#     data_all_list.append(unicode)

# 各表总数量
# print(len(data_ban_list))
# print(len(data_all_list))


# right=0
# wrong_count = 0
# wrong_list = []
# same = [x for x in data_ban_list if x in data_all_list] #找两表相同的内容
# no_same_in_Unicode = [x for x in data_all_list if x not in same]#在allUnicode中和banUnicodee不同的数据
# no_same_in_banUnicode = [x for x in data_all_list if x not in data_ban_list] #在banUnicode中找和allUnicode不同的数据

# print(same)
# print(len(same))
# print(len(no_same_in_Unicode))
# print(len(no_same_in_banUnicode))

# banUnicode中不同的
# for i in no_same_in_banUnicode:
#     print(i)

# allUnicode中不同的
# print("对比==========================")
# for i in no_same_in_Unicode:
#     print(i)
